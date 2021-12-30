# -*- coding: utf-8 -*-
"""
Created on Thu Dec  9 15:30:00 2021

@author: 20202619
"""


def Matlab_Calc(amount_peoples,weight,carbs,eat_time):

    import matlab.engine
    import openpyxl
    import csv

    def ExcelReader(filename):
        if 'xlsx' not in filename:
            print('none')
        else:
            workbook = openpyxl.load_workbook(filename)
            return(workbook)

    #data workage values
    workbook = ExcelReader(r"C:\Users\jeroe\OneDrive - TU Eindhoven\Vakken\Digital Twin 2\Matlab EDES python\Data_glucose_Val.xlsx")
    workbook_b = ExcelReader(r"C:\Users\jeroe\OneDrive - TU Eindhoven\Vakken\Digital Twin 2\Matlab EDES python\Data_insulin_Val.xlsx")
    workbook_ba = workbook_b.active
    workbook_a = workbook.active
    max_val_row = workbook_a.max_row
    Column_val = 2
    Person_val = 1
    list_people_pre_ad = []
    list_people_pre_glu_ins = []

    #matlab values
    weight = float(weight)
    carbs = float(carbs)
    eat_time = float(eat_time)
    determine_str = ""

    for i in range(2,(max_val_row+1)):
        cell = workbook_a.cell(row = i, column = Column_val)
        cell_1 = workbook_a.cell(row = i, column = Person_val) 
        cell_val = float(cell.value)
        cell_1_val = cell_1.value
        if cell_val > 5.5 and cell_val <= 6.9:
            list_people_pre_ad.append(cell_1_val)
        
    for n in range(2, (max_val_row+1)):
        for v in range(2,11):
            base_cell = workbook_a.cell(row = n, column = 1)
            base_cell_val = base_cell.value
            if base_cell_val in list_people_pre_ad:
                cell_pre = workbook_a.cell(row = n, column = v)
                cell_val_pre = cell_pre.value
                if cell_val_pre == None and base_cell_val in list_people_pre_ad:
                    list_people_pre_ad.remove(base_cell_val)
            base_cell_in = workbook_ba.cell(row = n, column = 1)
            base_cell_val_in = base_cell_in.value
            if base_cell_val_in in list_people_pre_ad:
                cell_pre_b = workbook_ba.cell(row = n, column = v)
                cell_val_pre_b = cell_pre_b.value
                if cell_val_pre_b == None and cell_val_pre_b in list_people_pre_ad:
                    list_people_pre_ad.remove(cell_val_pre_b)
    
    for k in range(2,(max_val_row+1)):
        base_cell = workbook_a.cell(row = k, column = 1)
        base_cell_val = base_cell.value
        if base_cell_val in list_people_pre_ad:
            for i in range(2,7):
                cell_0_pre = workbook_a.cell(row = k, column = i)
                cell_val_0 = cell_0_pre.value
                list_people_pre_glu_ins.append(str(cell_val_0))
        base_cell_in = workbook_ba.cell(row = k, column = 1)
        base_cell_val_in = base_cell_in.value
        if base_cell_val_in in list_people_pre_ad:
            for z in range(2,7):
                cell_0_pre_in = workbook_ba.cell(row = k, column = z)
                cell_val_0_in = cell_0_pre_in.value
                list_people_pre_glu_ins.append(str(cell_val_0_in))      
          
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data_Glu_Ins_pre"
    cell_1 = ws1.cell(column=1, row=1)
    cell_1.value = "glucose0,glucose15,glucose30,glucose60,glucose120,insulin0,insulin15,insulin30,insulin60,insulin120"
    
    for row in range(2,(len(list_people_pre_ad)+2)):
        list_temp = []
        cell = ws1.cell(column=1, row=row)
        for m in range(10):
            list_temp.append(list_people_pre_glu_ins.pop(0))
        cell.value = (','.join(list_temp))
    
    with open(r"C:\Users\jeroe\OneDrive - TU Eindhoven\Vakken\Digital Twin 2\Matlab EDES python\Data_Glu_Ins_pre.csv", 'w', newline='', encoding='utf-8') as csvfile:
        for i in range(1,(amount_peoples+2)):
            list_excel_rows = []
            cell_excel_rows = ws1.cell(column=1, row=i)
            list_excel_rows.append(cell_excel_rows.value)
            excelwriter = csv.writer(csvfile, delimiter=' ', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            excelwriter.writerow(list_excel_rows)
    wb.save(filename = r"C:\Users\jeroe\OneDrive - TU Eindhoven\Vakken\Digital Twin 2\Matlab EDES python\Data_Glu_Ins_pre.xlsx")
    
    eng = matlab.engine.start_matlab()
    peoples_list = eng.run_model_fun(weight,carbs,eat_time)
    person_1_list = list(peoples_list[0])
    max_val_people_1 = max(person_1_list)
    if max_val_people_1 >= 9.5:
        determine_str = "NO GOOD NO EAT"
    else:
        determine_str = "YES GOOD YES EAT"

    return(determine_str)

